import json
import openpyxl
from pathlib import Path

def leer_stakeholders(ruta: str = "config/stakeholders.json") -> dict:
    """Lee el archivo de configuración de stakeholders."""
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)

def leer_contexto(ruta: str = "config/contexto_proyecto.md") -> str:
    """Lee el contexto del proyecto en markdown."""
    with open(ruta, "r", encoding="utf-8") as f:
        return f.read()

def leer_proyecto(ruta: str = "data/lakehouse_project.xlsx") -> dict:
    """Lee el estado del proyecto desde el xlsx."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    resultado = {}

    for nombre_sheet in wb.sheetnames:
        ws = wb[nombre_sheet]
        filas = []
        headers = []

        for i, fila in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h) if h else "" for h in fila]
            else:
                if any(fila):
                    fila_dict = dict(zip(headers, fila))
                    filas.append(fila_dict)

        resultado[nombre_sheet] = filas

    return resultado